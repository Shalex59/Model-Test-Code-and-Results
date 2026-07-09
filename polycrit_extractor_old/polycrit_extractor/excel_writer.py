"""Writes extracted rows to a formatted .xlsx, in the same spirit as the
manually-built PolyCrit Excel files (Arial font, frozen header, autofilter,
color-coded by classification, plus a methodology/notes sheet)."""

from __future__ import annotations

from typing import List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .schema import EXCEL_COLUMNS

FONT = "Arial"

HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
CELL_FONT = Font(name=FONT, size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Row fill by classification (reached_lccc), applied unless overridden by a flag
STATUS_FILL = {
    "yes": PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA"),           # green: true LCCC
    "borderline/near-critical": PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC"),  # yellow
    "no": PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6"),            # orange: non-critical
    "unclear": PatternFill("solid", start_color="EDEDED", end_color="EDEDED"),       # gray
}
CONTRADICTION_FILL = PatternFill("solid", start_color="F8CBAD", end_color="F8CBAD")  # red-orange
INVALID_FONT_COLOR = "999999"

DEFAULT_WIDTHS = {
    "paper": 18, "doi": 16, "publication_year": 10, "corresponding_author": 18,
    "email": 20, "physical_address": 26, "sample_id": 14, "analyte_polymer": 28,
    "critical_component": 20, "architecture_note": 26, "condition_basis": 28,
    "column_name": 16, "stationary_phase_chemistry": 16, "column_mode": 12,
    "pore_size": 12, "column_dimensions": 16, "mobile_phase_solvents": 16,
    "mobile_phase_ratio": 14, "mobile_phase_ratio_units": 14, "aqueous_ph": 10,
    "aqueous_salt_added": 12, "aqueous_salt_type": 12, "aqueous_salt_concentration": 14,
    "temperature_c": 10, "flow_rate": 10, "detector": 14, "molar_mass_info": 26,
    "value_type": 18, "measured_value": 14, "reached_lccc": 16,
    "interaction_parameter_c": 14, "c_meaning": 32, "valid": 16, "evidence": 30,
    "notes": 26, "source_chunk": 12, "flag": 16, "flag_detail": 30,
}


def _try_number(val: str):
    """Best-effort numeric coercion for measured_value / temperature so the
    Excel cell is a real number (sortable/chartable) when possible, without
    losing non-numeric values like 'irreversible adsorption'."""
    if val is None:
        return ""
    s = str(val).strip()
    if s == "":
        return ""
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return s


NUMERIC_FIELDS = {"measured_value", "temperature_c", "publication_year"}


def write_excel(
    rows: List[dict],
    output_path: str,
    *,
    sheet_name: str = "extracted_data",
    methodology_notes: Optional[List[str]] = None,
) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    for j, (key, label) in enumerate(EXCEL_COLUMNS, start=1):
        c = ws.cell(row=1, column=j, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORDER
    ws.row_dimensions[1].height = 32

    for i, row in enumerate(rows, start=2):
        status_key = str(row.get("reached_lccc", "")).strip().lower()
        fill = STATUS_FILL.get(status_key)
        if str(row.get("flag", "")).upper() == "CONTRADICTION":
            fill = CONTRADICTION_FILL
        is_invalid = str(row.get("valid", "")).strip().lower().startswith("no")

        for j, (key, label) in enumerate(EXCEL_COLUMNS, start=1):
            val = row.get(key, "")
            val = _try_number(val) if key in NUMERIC_FIELDS else val
            c = ws.cell(row=i, column=j, value=val)
            c.font = Font(name=FONT, size=10, color=INVALID_FONT_COLOR, italic=True) if is_invalid else CELL_FONT
            c.alignment = WRAP
            c.border = BORDER
            if fill:
                c.fill = fill

    for j, (key, _label) in enumerate(EXCEL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = DEFAULT_WIDTHS.get(key, 14)

    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLUMNS))}{len(rows) + 1}"

    if methodology_notes:
        notes_ws = wb.create_sheet("Extraction Log")
        notes_ws.column_dimensions["A"].width = 110
        for i, line in enumerate(methodology_notes, start=1):
            c = notes_ws.cell(row=i, column=1, value=line)
            c.font = Font(name=FONT, bold=(i == 1), size=(12 if i == 1 else 10))
            c.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output_path)
    return output_path
